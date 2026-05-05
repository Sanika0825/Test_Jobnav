import openpyxl
class ExcelUtils:
    def __init__(self, file_path, sheet_name="Sheet1"):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def get_data(self):
        """Reads data from the Excel file and returns a list of dictionaries."""
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        
        data = []
        # Get headers from the first row
        headers = [cell.value for cell in sheet[1]]
        
        # Iterate over the rows starting from the second row
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): # Skip empty rows
                continue
            
            row_data = dict(zip(headers, row))
            row_data['row_index'] = row_index  # Keep track of row for updating later
            data.append(row_data)
            
        wb.close()
        return data

    def update_status(self, row_index, status_col_name, status_value):
        """Updates the status of a specific row."""
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        
        # Find the column index for the status column
        status_col_index = None
        for col_index, cell in enumerate(sheet[1], start=1):
            if cell.value == status_col_name:
                status_col_index = col_index
                break
                
        if status_col_index:
            sheet.cell(row=row_index, column=status_col_index).value = status_value
            wb.save(self.file_path)
        # this is a new comment to test git
        wb.close()
